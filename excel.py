import os
import sys

import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame

from ai import AI
from util.log import logger


class Excel(object):
    # EXCEL
    file_path: str = None
    dir_path: str = None
    dir_file_list: list[str] = list()
    data: DataFrame = None
    result: dict = dict()
    ai: AI = None
    progress: int = 0

    def __init__(self, file_path: str, dir_path: str) -> None:
        if not os.path.isfile(file_path):
            logger.error(f"No file selected: {file_path}")
            sys.exit(1)
        self.file_path = file_path
        if not os.path.isdir(dir_path):
            logger.error(f"No dir selected: {dir_path}")
            sys.exit(1)
        self.dir_path = dir_path
        self.dir_file_list = os.listdir(dir_path)
        if len(self.dir_file_list) == 0:
            logger.error(f"Dir :{self.dir_path} contains 0 file")
            sys.exit(1)
        # 读取Excel文件
        read_data = pd.read_excel(self.file_path)
        logger.debug(f"File columns: {read_data.columns}")
        self.data = read_data.copy()
        self.validate()
        self.ai = AI()

    def execute(self, progress_callback=None):
        # 找到多部门参与的项目
        # filter_project_codes = self.data.query('主责交付部门 != 涉及产品部门')['项目编码'].drop_duplicates()
        # 统计每个项目编码对应的产品部门
        project_department_info = self.data.groupby('项目编码')['涉及产品部门'].apply(list).to_dict()
        # 筛选出项目编码出现多次且涉及不同产品部门的项目编码
        filtered_projects = {k: v for k, v in project_department_info.items() if len(set(v)) > 1}
        logger.debug(filtered_projects)
        # 筛选出原始数据中符合条件的行
        filtered_df = self.data[self.data['项目编码'].isin(filtered_projects)]
        logger.debug(filtered_df)
        # 去重并提取项目编码
        unique_project_codes = filtered_projects.keys()
        logger.info(unique_project_codes)
        cur = 0
        total = len(unique_project_codes)
        for project_code in unique_project_codes:
            found = False
            for file in self.dir_file_list:
                if file.lower().find(project_code.lower()) != -1:
                    found = True
                    file_path = self.dir_path + "\\" + file
                    self.result[project_code] = self.dir_path + "\\" + file
                    logger.info(f"正在分析项目 {project_code}，部门 {filtered_projects.get(project_code)}, 文件 {file_path}")
                    try:
                        response = self.ai.analysis(file_path, depts=filtered_projects.get(project_code))
                        for dept in response:
                            self.data.loc[(self.data['项目编码'] == project_code) & (self.data['涉及产品部门'] == dept),
                                          '各部门自研成本'] = response.get(dept).get("total")
                            self.data.loc[(self.data['项目编码'] == project_code) & (self.data['涉及产品部门'] == dept),
                                          '各部门自研成本占比'] = response.get(dept).get("rate")

                    except Exception as e:
                        self.data.loc[(self.data['项目编码'] == project_code),
                                      '各部门自研成本'] = 'AI识别异常'
                        logger.error(e)
                    break
            cur = cur + 1
            self.progress = int(cur / total * 100)
            if progress_callback:
                progress_callback(self.progress)
            logger.info(f"progress is {self.progress}%")
            if not found:
                logger.warning(f"没找到项目 {project_code} 对应的自研成本评估明细")
                self.data.loc[(self.data['项目编码'] == project_code),
                              '各部门自研成本'] = '没找到自研成本评估明细'
        # 将结果反写回本文件
        # 获取原始文件的目录和文件名
        directory, filename = os.path.split(self.file_path)
        # 构造新的文件名，原始文件名加上"_分析后"
        if filename.endswith('.xls'):
            temp_filename = filename.replace('.xls', '_临时文件.xls')
            new_filename = filename.replace('.xls', '_分析后.xls')
        else:
            temp_filename = filename.replace('.xlsx', '_临时文件.xlsx')
            new_filename = filename.replace('.xlsx', '_分析后.xlsx')
        # 完整的新文件路径
        temp_file_path = os.path.join(directory, temp_filename)
        new_file_path = os.path.join(directory, new_filename)
        # 将处理后的DataFrame写入到新的Excel文件
        self.data.to_excel(temp_file_path, index=False)

        # 获取目标工作表
        original_wb = load_workbook(filename=self.file_path)
        target_sheet_name = 'Sheet1'
        # 如果Excel文件只有一个工作表，获取其名称
        if len(original_wb.sheetnames) == 1:
            target_sheet_name = original_wb.sheetnames[0]
            logger.debug(f"The single sheet name is: {target_sheet_name}")
        else:
            logger.error("The Excel file contains more than one sheet.")
            raise Exception("The Excel file contains more than one sheet.")
        target_ws = original_wb[target_sheet_name]

        # 获取列索引
        col_indices = {cell.value: idx for idx, cell in enumerate(target_ws[1], 1)}
        income_col_idx = col_indices.get('各部门自研成本')
        ratio_col_idx = col_indices.get('各部门自研成本占比')

        if income_col_idx is None or ratio_col_idx is None:
            logger.error("未找到指定的列名")
            raise Exception("未找到指定的列名")

        # 获取临时表中的数据
        temp_data = pd.read_excel(temp_file_path, usecols=['各部门自研成本', '各部门自研成本占比'])

        # 将临时表中的数据写入到原始表中
        for row_idx, row in temp_data.iterrows():
            target_ws.cell(row=row_idx + 2, column=income_col_idx, value=row['各部门自研成本'])
            target_ws.cell(row=row_idx + 2, column=ratio_col_idx, value=row['各部门自研成本占比'])

        # 保存修改后的原始Excel文件
        original_wb.save(filename=new_file_path)
        os.remove(temp_file_path)

    def validate(self):
        if "涉及产品部门" not in self.data.columns:
            logger.error(f"文件中应该包含: 涉及产品部门")
            sys.exit(1)
        if "项目编码" not in self.data.columns:
            logger.error(f"文件中应该包含: 项目编码")
            sys.exit(1)


if __name__ == '__main__':
    file_path = 'C:\\Users\\linli\\Desktop\\在岗创新\\ICT项目收入明细表（2024）-7.xlsx'
    directory_path = 'C:\\Users\\linli\\Desktop\\在岗创新\\1-7月涉及多部门的自研成本测算表-修改'
    excel = Excel(file_path, directory_path)
    excel.execute()
